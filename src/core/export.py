"""
Export functionality for browser history data
"""
import json
import csv
import html
from pathlib import Path
from typing import List
from dataclasses import asdict
from datetime import datetime

from .models import HistoryEntry, Download, Cookie, Bookmark
from ..utils.security import validate_export_path, escape_html


class DataExporter:
    """Export browser history data to various formats"""

    @staticmethod
    def export_to_csv(entries: List, output_path: str, timezone: str = "UTC"):
        """
        Export entries to CSV

        Args:
            entries: List of entries
            output_path: Output file path
            timezone: Timezone for date formatting
        """
        from .timezone_utils import TimezoneConverter

        # Security: Validate output path
        if not validate_export_path(output_path, allowed_extensions=('.csv',)):
            raise ValueError(f"Invalid or unsafe output path: {output_path}")

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            if not entries:
                return

            # Determine fields based on first entry type
            first_entry = entries[0]

            if isinstance(first_entry, HistoryEntry):
                fieldnames = [
                    'id', 'url', 'title', 'visit_time', 'visit_count',
                    'browser', 'domain', 'typed_count', 'last_visit_time',
                    'hidden', 'notes', 'bookmarked', 'source_file', 'source_file_hash'
                ]
            elif isinstance(first_entry, Download):
                fieldnames = [
                    'id', 'url', 'target_path', 'start_time', 'end_time',
                    'total_bytes', 'received_bytes', 'state', 'danger_type',
                    'browser', 'mime_type', 'source_file'
                ]
            elif isinstance(first_entry, Cookie):
                fieldnames = [
                    'host_key', 'name', 'value', 'path', 'creation_utc',
                    'expires_utc', 'last_access_utc', 'is_secure', 'is_httponly',
                    'has_expires', 'is_persistent', 'browser', 'source_file'
                ]
            elif isinstance(first_entry, Bookmark):
                fieldnames = [
                    'id', 'url', 'title', 'date_added', 'parent_folder',
                    'browser', 'source_file'
                ]
            else:
                # Generic export
                fieldnames = list(asdict(first_entry).keys())

            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for entry in entries:
                row_dict = asdict(entry)

                # Convert datetime objects to strings with timezone
                for key, value in row_dict.items():
                    if isinstance(value, datetime):
                        converted = TimezoneConverter.convert_timezone(value, timezone)
                        row_dict[key] = TimezoneConverter.format_datetime(converted)
                    elif isinstance(value, list):
                        row_dict[key] = ', '.join(str(v) for v in value)
                    elif isinstance(value, dict):
                        row_dict[key] = json.dumps(value)

                # Filter to only include fieldnames
                filtered_dict = {k: row_dict.get(k, '') for k in fieldnames}
                writer.writerow(filtered_dict)

    @staticmethod
    def export_to_json(entries: List, output_path: str, timezone: str = "UTC", pretty: bool = True):
        """
        Export entries to JSON

        Args:
            entries: List of entries
            output_path: Output file path
            timezone: Timezone for date formatting
            pretty: Pretty print JSON
        """
        from .timezone_utils import TimezoneConverter

        # Security: Validate output path
        if not validate_export_path(output_path, allowed_extensions=('.json',)):
            raise ValueError(f"Invalid or unsafe output path: {output_path}")

        data = []
        for entry in entries:
            entry_dict = asdict(entry)

            # Convert datetime objects to strings
            for key, value in entry_dict.items():
                if isinstance(value, datetime):
                    converted = TimezoneConverter.convert_timezone(value, timezone)
                    entry_dict[key] = TimezoneConverter.format_datetime(converted, fmt="%Y-%m-%d %H:%M:%S %Z")

            data.append(entry_dict)

        with open(output_path, 'w', encoding='utf-8') as f:
            if pretty:
                json.dump(data, f, indent=2, ensure_ascii=False)
            else:
                json.dump(data, f, ensure_ascii=False)

    @staticmethod
    def export_to_excel(entries: List, output_path: str, timezone: str = "UTC"):
        """
        Export entries to Excel

        Args:
            entries: List of entries
            output_path: Output file path
            timezone: Timezone for date formatting
        """
        import pandas as pd
        from .timezone_utils import TimezoneConverter

        # Security: Validate output path
        if not validate_export_path(output_path, allowed_extensions=('.xlsx', '.xls')):
            raise ValueError(f"Invalid or unsafe output path: {output_path}")

        if not entries:
            return

        # Convert to dictionaries
        data = []
        for entry in entries:
            entry_dict = asdict(entry)

            # Convert datetime objects to strings
            for key, value in entry_dict.items():
                if isinstance(value, datetime):
                    converted = TimezoneConverter.convert_timezone(value, timezone)
                    entry_dict[key] = TimezoneConverter.format_datetime(converted, fmt="%Y-%m-%d %H:%M:%S %Z")
                elif isinstance(value, list):
                    entry_dict[key] = ', '.join(str(v) for v in value)
                elif isinstance(value, dict):
                    entry_dict[key] = json.dumps(value)

            data.append(entry_dict)

        # Create DataFrame
        df = pd.DataFrame(data)

        # Export to Excel
        df.to_excel(output_path, index=False, engine='openpyxl')

    @staticmethod
    def export_to_html(entries: List, output_path: str, timezone: str = "UTC", title: str = "Browser History Report"):
        """
        Export entries to HTML

        Args:
            entries: List of entries
            output_path: Output file path
            timezone: Timezone for date formatting
            title: Report title
        """
        from .timezone_utils import TimezoneConverter

        # Security: Validate output path
        if not validate_export_path(output_path, allowed_extensions=('.html', '.htm')):
            raise ValueError(f"Invalid or unsafe output path: {output_path}")

        # Security: Escape title and timezone to prevent XSS
        title = escape_html(title)
        timezone_escaped = escape_html(timezone)

        html_template = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{title}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        h1 {{
            color: #333;
            border-bottom: 3px solid #4CAF50;
            padding-bottom: 10px;
        }}
        .metadata {{
            background-color: #fff;
            padding: 15px;
            margin-bottom: 20px;
            border-radius: 5px;
            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            background-color: #fff;
            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        }}
        th {{
            background-color: #4CAF50;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .url {{
            color: #1a0dab;
            word-break: break-all;
        }}
        .browser {{
            display: inline-block;
            padding: 3px 8px;
            border-radius: 3px;
            font-size: 12px;
            font-weight: bold;
        }}
        .chrome {{ background-color: #ffd966; }}
        .firefox {{ background-color: #ff9966; }}
        .edge {{ background-color: #99ccff; }}
    </style>
</head>
<body>
    <h1>{title}</h1>
    <div class="metadata">
        <strong>Generated:</strong> {generated_time}<br>
        <strong>Timezone:</strong> {timezone}<br>
        <strong>Total Entries:</strong> {total_entries}
    </div>
    <table>
        <thead>
            <tr>
                {headers}
            </tr>
        </thead>
        <tbody>
            {rows}
        </tbody>
    </table>
</body>
</html>
"""

        if not entries:
            html = html_template.format(
                title=title,
                generated_time=escape_html(datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                timezone=timezone_escaped,
                total_entries=0,
                headers="<th>No data</th>",
                rows="<tr><td>No entries found</td></tr>"
            )
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html)
            return

        # Generate headers based on entry type
        first_entry = entries[0]
        if isinstance(first_entry, HistoryEntry):
            headers = ['Browser', 'Visit Time', 'URL', 'Title', 'Visit Count', 'Domain']
            # Security: Headers are constants, but escape for consistency
            header_html = ''.join(f'<th>{escape_html(h)}</th>' for h in headers)

            # Generate rows
            rows_html = []
            for entry in entries:
                visit_time = TimezoneConverter.convert_timezone(entry.visit_time, timezone)
                time_str = TimezoneConverter.format_datetime(visit_time, fmt="%Y-%m-%d %H:%M:%S")

                # Security: Escape all user-controlled data to prevent XSS
                browser_escaped = escape_html(entry.browser)
                browser_class = entry.browser.lower()  # CSS class name, limited set
                url_escaped = escape_html(entry.url)
                title_escaped = escape_html(entry.title or 'N/A')
                domain_escaped = escape_html(entry.domain)

                row = f"""
                <tr>
                    <td><span class="browser {browser_class}">{browser_escaped}</span></td>
                    <td>{escape_html(time_str)}</td>
                    <td class="url">{url_escaped}</td>
                    <td>{title_escaped}</td>
                    <td>{entry.visit_count}</td>
                    <td>{domain_escaped}</td>
                </tr>
                """
                rows_html.append(row)

        else:
            # Generic export for other types
            entry_dict = asdict(first_entry)
            headers = list(entry_dict.keys())[:10]  # Limit to 10 columns for readability
            # Security: Escape header names
            header_html = ''.join(f'<th>{escape_html(str(h))}</th>' for h in headers)

            rows_html = []
            for entry in entries:
                entry_dict = asdict(entry)
                cells = []
                for header in headers:
                    value = entry_dict.get(header, '')
                    if isinstance(value, datetime):
                        converted = TimezoneConverter.convert_timezone(value, timezone)
                        value = TimezoneConverter.format_datetime(converted, fmt="%Y-%m-%d %H:%M:%S")
                    # Security: Escape all values to prevent XSS
                    cells.append(f'<td>{escape_html(str(value))}</td>')
                rows_html.append('<tr>' + ''.join(cells) + '</tr>')

        html = html_template.format(
            title=title,
            generated_time=escape_html(datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            timezone=timezone_escaped,
            total_entries=len(entries),
            headers=header_html,
            rows='\n'.join(rows_html)
        )

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)

    @staticmethod
    def export_generic_to_csv(data: List[dict], output_path: str):
        """
        Export generic dictionary data to CSV

        Args:
            data: List of dictionaries
            output_path: Output file path
        """
        # Security: Validate output path
        if not validate_export_path(output_path, allowed_extensions=('.csv',)):
            raise ValueError(f"Invalid or unsafe output path: {output_path}")

        if not data:
            # Create empty file
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                pass
            return

        # Get all keys from first row
        fieldnames = list(data[0].keys())

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)

    @staticmethod
    def export_generic_to_json(data: List[dict], output_path: str):
        """
        Export generic dictionary data to JSON

        Args:
            data: List of dictionaries
            output_path: Output file path
        """
        # Security: Validate output path
        if not validate_export_path(output_path, allowed_extensions=('.json',)):
            raise ValueError(f"Invalid or unsafe output path: {output_path}")

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)

    @staticmethod
    def export_generic_to_excel(data: List[dict], output_path: str):
        """
        Export generic dictionary data to Excel

        Args:
            data: List of dictionaries
            output_path: Output file path
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        # Security: Validate output path
        if not validate_export_path(output_path, allowed_extensions=('.xlsx',)):
            raise ValueError(f"Invalid or unsafe output path: {output_path}")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        if not data:
            wb.save(output_path)
            return

        # Get headers from first row
        headers = list(data[0].keys())

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data rows
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
